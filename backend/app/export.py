"""Report export renderers (Phase 9.1) — CSV / Excel / PDF from a generic snapshot.

The snapshot is the dict returned by calculations.report_* : {type,title,period,kpis,sections}.
Money/date cells are pre-formatted strings; this module only renders `text`.
PDF needs a Cyrillic-capable TTF (reportlab built-ins don't render Cyrillic) — see _register_fonts.
Spec: docs/superpowers/specs/2026-06-28-phase9-reports-design.md (R9/R10).
"""
from __future__ import annotations

import io
import os


def _cell_text(cell) -> str:
    """Plain-text rendering of a cell (str or styled object)."""
    if isinstance(cell, dict):
        if cell.get("kind") == "claim":
            code = cell.get("code") or "—"
            title = cell.get("title") or ""
            return f"{code} {title}".strip()
        return cell.get("text") or ""
    return str(cell)


# ---------------------------------------------------------------- CSV

def render_csv(snap: dict) -> str:
    buf = io.StringIO()
    buf.write("﻿")  # UTF-8 BOM — Excel opens Cyrillic correctly
    import csv
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    w.writerow([snap["title"]])
    if snap.get("period"):
        p = snap["period"]
        w.writerow([p["label"], p.get("from", ""), p.get("to", "")])
    for kpi in snap["kpis"]:
        w.writerow([kpi["label"], kpi["value"]])
    for sec in snap["sections"]:
        if sec.get("title"):
            w.writerow([])
            w.writerow([sec["title"]])
        w.writerow([c["label"] for c in sec["columns"]])
        for row in sec["rows"]:
            w.writerow([_cell_text(c) for c in row])
        if sec.get("footer"):
            w.writerow([_cell_text(c) for c in sec["footer"]])
    return buf.getvalue()


# ---------------------------------------------------------------- Excel

def render_excel(snap: dict) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"
    r = 1
    ws.cell(r, 1, snap["title"]).font = Font(bold=True, size=14)
    r += 1
    if snap.get("period"):
        p = snap["period"]
        ws.cell(r, 1, p["label"]); ws.cell(r, 2, p.get("from", "")); ws.cell(r, 3, p.get("to", ""))
        r += 1
    for kpi in snap["kpis"]:
        ws.cell(r, 1, kpi["label"]); ws.cell(r, 2, kpi["value"])
        r += 1
    for sec in snap["sections"]:
        if sec.get("title"):
            ws.cell(r, 1, sec["title"]).font = Font(bold=True)
            r += 1
        for j, col in enumerate(sec["columns"], 1):
            ws.cell(r, j, col["label"]).font = Font(bold=True)
        r += 1
        for row in sec["rows"]:
            for j, c in enumerate(row, 1):
                ws.cell(r, j, _cell_text(c))
            r += 1
        if sec.get("footer"):
            for j, c in enumerate(sec["footer"], 1):
                ws.cell(r, j, _cell_text(c)).font = Font(bold=True)
            r += 1
        r += 1
    for col in ws.columns:
        maxlen = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(maxlen + 2, 60)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------- PDF

_FONT_CACHE = None


def _register_fonts():
    """Return (normal_name, bold_name) of a Cyrillic-capable TTF registered with reportlab.

    Priority: bundled backend/app/fonts/DejaVuSans*.ttf → Windows arial*.ttf → error.
    """
    global _FONT_CACHE
    if _FONT_CACHE:
        return _FONT_CACHE
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    here = os.path.dirname(__file__)
    candidates = [
        (os.path.join(here, "fonts", "DejaVuSans.ttf"),
         os.path.join(here, "fonts", "DejaVuSans-Bold.ttf"), "DJ", "DJB"),
        (r"C:\Windows\Fonts\arial.ttf", r"C:\Windows\Fonts\arialbd.ttf", "AR", "ARB"),
        (r"C:\Windows\Fonts\consola.ttf", r"C:\Windows\Fonts\consolab.ttf", "CON", "CONB"),
    ]
    for normal, bold, nname, bname in candidates:
        if os.path.exists(normal):
            try:
                pdfmetrics.registerFont(TTFont(nname, normal))
                bold_name = nname
                if os.path.exists(bold):
                    pdfmetrics.registerFont(TTFont(bname, bold))
                    bold_name = bname
                _FONT_CACHE = (nname, bold_name)
                return _FONT_CACHE
            except Exception:
                continue
    raise RuntimeError(
        "No Cyrillic-capable TTF found for PDF export. "
        "Bundle DejaVuSans.ttf in backend/app/fonts/ or run on Windows."
    )


def _escape(s: str) -> str:
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def render_pdf(snap: dict) -> io.BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    font_normal, font_bold = _register_fonts()
    s_title = ParagraphStyle("t", fontName=font_bold, fontSize=14, leading=18)
    s_meta = ParagraphStyle("m", fontName=font_normal, fontSize=9, leading=12, textColor=colors.HexColor("#666666"))
    s_kpi = ParagraphStyle("k", fontName=font_normal, fontSize=9, leading=13)
    s_h = ParagraphStyle("h", fontName=font_bold, fontSize=10, leading=14, spaceBefore=6)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=12 * mm, leftMargin=12 * mm, rightMargin=12 * mm, bottomMargin=12 * mm)
    elems = [Paragraph(_escape(snap["title"]), s_title)]
    if snap.get("period"):
        p = snap["period"]
        elems.append(Paragraph(
            _escape(f"{p['label']}: {p.get('from', '')} — {p.get('to', '')}"), s_meta))
    elems.append(Paragraph(
        _escape(" · ".join(f"{k['label']}: {k['value']}" for k in snap["kpis"])), s_kpi))
    elems.append(Spacer(1, 6))

    page_w = A4[0] - 24 * mm
    for sec in snap["sections"]:
        if sec.get("title"):
            elems.append(Paragraph(_escape(sec["title"]), s_h))
        data = [[_escape(c["label"]) for c in sec["columns"]]]
        aligns = [c.get("align") for c in sec["columns"]]
        for row in sec["rows"]:
            data.append([_escape(_cell_text(c)) for c in row])
        if sec.get("footer"):
            data.append([_escape(_cell_text(c)) for c in sec["footer"]])
        ncol = len(sec["columns"])
        widths = [page_w / ncol] * ncol
        t = Table(data, colWidths=widths, repeatRows=1)
        style = [
            ("FONT", (0, 0), (-1, -1), font_normal, 8),
            ("FONT", (0, 0), (-1, 0), font_bold, 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef0f3")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for j, a in enumerate(aligns):
            if a == "right":
                style.append(("ALIGN", (j, 0), (j, -1), "RIGHT"))
        if sec.get("footer"):
            style.append(("FONT", (0, -1), (-1, -1), font_bold, 8))
            style.append(("LINEABOVE", (0, -1), (-1, -1), 0.8, colors.black))
        t.setStyle(TableStyle(style))
        elems.append(t)
        elems.append(Spacer(1, 8))

    doc.build(elems)
    buf.seek(0)
    return buf
